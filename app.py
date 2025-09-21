import streamlit as st
import pandas as pd
import plotly.express as px
import numpy as np
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Supplier Capacity Dashboard", layout="wide")
st.title("📊 Supplier Capacity Dashboard")

# ================= Upload file =================
uploaded_file = st.file_uploader("Upload Excel Input", type=["xlsx"])
if uploaded_file:
    # Đọc dữ liệu
    capacity_df = pd.read_excel(uploaded_file, sheet_name="Capacity")
    demand_df = pd.read_excel(uploaded_file, sheet_name="Demand")

    # ===== Tính Capacity =====
    capacity_df["Capacity"] = (
        capacity_df["Lines"] *
        capacity_df["HoursPerDay"] *
        capacity_df["OutputPerHourPerLine"] *
        capacity_df["WorkingDays"]
    )

    # ===== Demand reshape (Month đã ở dạng YYYY-mm) =====
    demand_long = demand_df.melt(
        id_vars=["Vendor","Item","Process"],
        var_name="Month", value_name="Demand"
    )
    demand_long["Month"] = pd.to_datetime(demand_long["Month"], format="%Y-%m")

    demand_sum = demand_long.groupby(["Vendor","Process","Month"])["Demand"].sum().reset_index()

    # Merge với capacity
    merged = demand_sum.merge(
        capacity_df[["Vendor","Process","Capacity"]],
        on=["Vendor","Process"], how="left"
    )
    # ✅ Fulfillment% = Capacity / Demand * 100
    merged["Fulfillment_%"] = (merged["Capacity"] / merged["Demand"] * 100).round(2)
    merged["Status"] = merged.apply(lambda r: "OK" if r["Fulfillment_%"] >= 100 else "Shortage", axis=1)

    # ===== Summary theo Vendor =====
    summary_vendor = merged.groupby(["Vendor","Month"]).agg({
        "Capacity":"min",   # bottleneck process
        "Demand":"sum"
    }).reset_index()
    summary_vendor["Fulfillment_%"] = (summary_vendor["Capacity"] / summary_vendor["Demand"] * 100).round(2)

    # ===== Summary toàn bộ =====
    summary_total = merged.groupby("Month").agg({
        "Capacity":"sum","Demand":"sum"
    }).reset_index()
    summary_total["Fulfillment_%"] = (summary_total["Capacity"] / summary_total["Demand"] * 100).round(2)

    # Sort theo thời gian
    summary_vendor = summary_vendor.sort_values(["Vendor","Month"])
    summary_total = summary_total.sort_values("Month")

    # ===== Slicer theo tháng =====
    months_available = sorted(summary_total["Month"].dt.strftime("%Y-%m").unique())
    months_selected = st.multiselect("📅 Chọn tháng:", months_available, default=months_available)

    if months_selected:
        summary_vendor = summary_vendor[summary_vendor["Month"].dt.strftime("%Y-%m").isin(months_selected)]
        summary_total = summary_total[summary_total["Month"].dt.strftime("%Y-%m").isin(months_selected)]

    # ===== Slicer chọn chế độ Fulfillment% =====
    st.subheader("🎛️ Bộ lọc Fulfillment%")
    mode = st.radio(
        "Chọn chế độ:",
        ["All", "Low (≤75%)", "Medium (75–85%)", "High (>85%)"]
    )

    filtered_vendor = summary_vendor.copy()
    if mode == "Low (≤75%)":
        filtered_vendor = filtered_vendor[filtered_vendor["Fulfillment_%"] <= 75]
    elif mode == "Medium (75–85%)":
        filtered_vendor = filtered_vendor[(filtered_vendor["Fulfillment_%"] > 75) & (filtered_vendor["Fulfillment_%"] <= 85)]
    elif mode == "High (>85%)":
        filtered_vendor = filtered_vendor[filtered_vendor["Fulfillment_%"] > 85]

    # ===== Hiển thị bảng =====
    st.subheader("🔎 Vendor Summary")
    st.dataframe(filtered_vendor)

    st.subheader("🌍 Total Supply Chain Summary")
    st.dataframe(summary_total)

    # ===== Chart Demand vs Capacity =====
    st.subheader("📊 Demand vs Capacity")
    vendor_selected = st.selectbox("Chọn Vendor", ["ALL"] + sorted(filtered_vendor["Vendor"].unique()))

    if vendor_selected == "ALL":
        chart_data = summary_total.melt(
            id_vars="Month", value_vars=["Demand","Capacity"], 
            var_name="Type", value_name="Value"
        )
        fig = px.bar(chart_data, x="Month", y="Value", color="Type", barmode="group",
                     title="Total Demand vs Capacity", text="Value")
    else:
        vendor_data = filtered_vendor[filtered_vendor["Vendor"] == vendor_selected]
        chart_data = vendor_data.melt(
            id_vars=["Month"], value_vars=["Demand","Capacity"], 
            var_name="Type", value_name="Value"
        )
        fig = px.bar(chart_data, x="Month", y="Value", color="Type", barmode="group",
                     title=f"Vendor {vendor_selected} - Demand vs Capacity", text="Value")

    fig.update_traces(textposition="outside")
    fig.update_layout(
        xaxis=dict(tickformat="%Y-%m", type="category"),
        uniformtext_minsize=8, uniformtext_mode="hide"
    )
    st.plotly_chart(fig, use_container_width=True)

    # ===== Heatmap Fulfillment% =====
    st.subheader("🔥 Fulfillment% Heatmap")

    heatmap_df = filtered_vendor.pivot_table(
        index="Vendor", 
        columns=filtered_vendor["Month"].dt.strftime("%Y-%m"),
        values="Fulfillment_%", aggfunc="mean"
    ).fillna(0)

    colorscale = [
        [0.0, "green"],   # ≤75%
        [0.75, "green"],
        [0.85, "yellow"], # 75–85%
        [1.0, "red"]      # >85%
    ]

    if not heatmap_df.empty:
        fig_heat = px.imshow(
            heatmap_df.values,
            x=heatmap_df.columns,
            y=heatmap_df.index,
            color_continuous_scale=colorscale,
            aspect="auto",
            labels=dict(x="Month", y="Vendor", color="Fulfillment%")
        )
        fig_heat.update_traces(
            text=np.round(heatmap_df.values, 1),
            texttemplate="%{text}",
            textfont=dict(size=10)
        )
        st.plotly_chart(fig_heat, use_container_width=True)
    else:
        st.info("⚠️ Không có dữ liệu để hiển thị heatmap với chế độ filter này.")

    # ===== Line Chart Fulfillment% =====
    st.subheader("📈 Fulfillment% Trend")

    if vendor_selected == "ALL":
        fig_line = px.line(summary_total, x="Month", y="Fulfillment_%",
                           title="Trend Fulfillment% - All Vendors")
    else:
        vendor_data = filtered_vendor[filtered_vendor["Vendor"] == vendor_selected]
        fig_line = px.line(vendor_data, x="Month", y="Fulfillment_%",
                           title=f"Trend Fulfillment% - {vendor_selected}")

    fig_line.update_traces(mode="lines+markers", line=dict(width=2))
    fig_line.update_layout(
        xaxis=dict(tickformat="%Y-%m", type="category"),
        yaxis=dict(title="Fulfillment %")
    )
    st.plotly_chart(fig_line, use_container_width=True)

    # ===== Xuất Excel (theo vendor chọn + filter) =====
    out_file = BytesIO()
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        capacity_df.to_excel(writer, sheet_name="Capacity_Input", index=False)
        demand_df.to_excel(writer, sheet_name="Demand_Input", index=False)

        if vendor_selected == "ALL":
            merged.to_excel(writer, sheet_name="Process_Result", index=False)
            filtered_vendor.to_excel(writer, sheet_name="Vendor_Summary", index=False)
            summary_total.to_excel(writer, sheet_name="Total_Summary", index=False)
        else:
            merged_vendor = merged[merged["Vendor"] == vendor_selected]
            summary_vendor_sel = filtered_vendor[filtered_vendor["Vendor"] == vendor_selected]

            merged_vendor.to_excel(writer, sheet_name=f"{vendor_selected}_Process", index=False)
            summary_vendor_sel.to_excel(writer, sheet_name=f"{vendor_selected}_Summary", index=False)

    # ===== Highlight Excel =====
    out_file.seek(0)
    wb = load_workbook(out_file)

    def format_sheet(ws):
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2
        for c in ws[1]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        # highlight theo Fulfillment_% nếu có
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if "Fulfillment" in str(ws[1][cell.column-1].value):
                    try:
                        val = float(cell.value)
                        if val < 100:
                            cell.fill = PatternFill("solid", fgColor="FFC7CE")  # đỏ nhạt
                        else:
                            cell.fill = PatternFill("solid", fgColor="C6EFCE")  # xanh nhạt
                    except:
                        pass

    for sheet in wb.sheetnames:
        format_sheet(wb[sheet])

    final_out = BytesIO()
    wb.save(final_out)

    file_name = (
        "Supplier_Capacity_Result_ALL.xlsx"
        if vendor_selected == "ALL"
        else f"Supplier_Capacity_Result_{vendor_selected}.xlsx"
    )

    st.download_button(
        label="⬇️ Download Result Excel",
        data=final_out.getvalue(),
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
